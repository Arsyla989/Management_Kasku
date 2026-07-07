import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from config import settings
from typing import List, Dict, Any

class ExcelService:
    @staticmethod
    def export_all_data(
        siswa_list: List[Any],
        pembayaran_list: List[Any],
        pengeluaran_list: List[Any],
        status_pembayaran: List[Dict[str, Any]],
        bulan_aktif: str
    ) -> str:
        wb = Workbook()
        
        # Style Definitions (Banking Blue Theme)
        font_title = Font(name="Segoe UI", size=16, bold=True, color="0F62FE")
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Segoe UI", size=10, bold=True)
        font_normal = Font(name="Segoe UI", size=10)
        
        fill_primary = PatternFill(start_color="0F62FE", end_color="0F62FE", fill_type="solid")
        fill_secondary = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        fill_success = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")
        fill_danger = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        # ----------------------------------------------------
        # SHEET 1: RINGKASAN LAPORAN
        # ----------------------------------------------------
        ws_summary = wb.active
        ws_summary.title = "Ringkasan Laporan"
        ws_summary.views.sheetView[0].showGridLines = True
        
        # Header Title
        ws_summary["B2"] = "LAPORAN RINGKASAN KAS KELAS (KASKU)"
        ws_summary["B2"].font = font_title
        ws_summary["B3"] = f"Dibuat pada: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_summary["B3"].font = Font(name="Segoe UI", size=9, italic=True, color="7F8C8D")
        
        # Calculate stats
        total_siswa = len(siswa_list)
        total_masuk = sum(p.nominal for p in pembayaran_list)
        total_keluar = sum(p.nominal for p in pengeluaran_list)
        saldo = total_masuk - total_keluar
        
        stats_labels = [
            ("Total Siswa", total_siswa, "Siswa"),
            ("Total Pemasukan Kas", total_masuk, "Rupiah"),
            ("Total Pengeluaran Kas", total_keluar, "Rupiah"),
            ("Saldo Kas Kelas", saldo, "Rupiah")
        ]
        
        # Write stats table
        ws_summary.cell(row=5, column=2, value="Parameter").font = font_header
        ws_summary.cell(row=5, column=2).fill = fill_primary
        ws_summary.cell(row=5, column=2).alignment = align_center
        ws_summary.cell(row=5, column=2).border = thin_border
        
        ws_summary.cell(row=5, column=3, value="Nilai").font = font_header
        ws_summary.cell(row=5, column=3).fill = fill_primary
        ws_summary.cell(row=5, column=3).alignment = align_center
        ws_summary.cell(row=5, column=3).border = thin_border
        
        ws_summary.row_dimensions[5].height = 24
        
        for idx, (label, val, val_type) in enumerate(stats_labels, 6):
            ws_summary.row_dimensions[idx].height = 20
            
            c_label = ws_summary.cell(row=idx, column=2, value=label)
            c_label.font = font_bold
            c_label.border = thin_border
            
            c_val = ws_summary.cell(row=idx, column=3, value=val)
            c_val.font = font_normal
            c_val.border = thin_border
            
            if val_type == "Rupiah":
                c_val.number_format = 'Rp #,##0'
                c_val.alignment = align_right
            else:
                c_val.alignment = align_center

        # ----------------------------------------------------
        # SHEET 2: DATA SISWA
        # ----------------------------------------------------
        ws_siswa = wb.create_sheet(title="Data Siswa")
        ws_siswa.views.sheetView[0].showGridLines = True
        
        ws_siswa.append(["DAFTAR SISWA KELAS"])
        ws_siswa["A1"].font = font_title
        ws_siswa.append([])
        
        headers_siswa = ["No ID", "NIS", "Nama Lengkap", "Kelas"]
        ws_siswa.append(headers_siswa)
        ws_siswa.row_dimensions[3].height = 24
        
        for col_idx, h in enumerate(headers_siswa, 1):
            cell = ws_siswa.cell(row=3, column=col_idx)
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = align_center
            cell.border = thin_border
            
        for s in siswa_list:
            ws_siswa.append([s.id, s.nis, s.nama, s.kelas])
            row_num = ws_siswa.max_row
            ws_siswa.row_dimensions[row_num].height = 18
            ws_siswa.cell(row=row_num, column=1).alignment = align_center
            ws_siswa.cell(row=row_num, column=2).alignment = align_center
            ws_siswa.cell(row=row_num, column=3).alignment = align_left
            ws_siswa.cell(row=row_num, column=4).alignment = align_center
            for col_idx in range(1, 5):
                ws_siswa.cell(row=row_num, column=col_idx).font = font_normal
                ws_siswa.cell(row=row_num, column=col_idx).border = thin_border

        # ----------------------------------------------------
        # SHEET 3: STATUS PEMBAYARAN BULAN AKTIF
        # ----------------------------------------------------
        ws_status = wb.create_sheet(title=f"Status Kas {bulan_aktif}")
        ws_status.views.sheetView[0].showGridLines = True
        
        ws_status.append([f"STATUS PEMBAYARAN KAS BULAN: {bulan_aktif.upper()}"])
        ws_status["A1"].font = font_title
        ws_status.append([f"Target Iuran Bulanan: Rp {settings.KAS_BULANAN_NOMINAL:,}"])
        ws_status["A2"].font = Font(name="Segoe UI", size=10, italic=True)
        ws_status.append([])
        
        headers_status = ["NIS", "Nama Siswa", "Kelas", "Total Bayar", "Status"]
        ws_status.append(headers_status)
        ws_status.row_dimensions[4].height = 24
        
        for col_idx, h in enumerate(headers_status, 1):
            cell = ws_status.cell(row=4, column=col_idx)
            cell.font = font_header
            cell.fill = fill_secondary
            cell.alignment = align_center
            cell.border = thin_border
            
        for st in status_pembayaran:
            ws_status.append([st["nis"], st["nama"], st["kelas"], st["total_bayar"], st["status"]])
            row_num = ws_status.max_row
            ws_status.row_dimensions[row_num].height = 18
            
            # Format numbers and styles
            ws_status.cell(row=row_num, column=1).alignment = align_center
            ws_status.cell(row=row_num, column=2).alignment = align_left
            ws_status.cell(row=row_num, column=3).alignment = align_center
            
            val_cell = ws_status.cell(row=row_num, column=4)
            val_cell.alignment = align_right
            val_cell.number_format = 'Rp #,##0'
            
            status_cell = ws_status.cell(row=row_num, column=5)
            status_cell.alignment = align_center
            status_cell.font = font_bold
            if st["status"] == "Lunas":
                status_cell.fill = fill_success
            else:
                status_cell.fill = fill_danger
                
            for col_idx in range(1, 6):
                ws_status.cell(row=row_num, column=col_idx).border = thin_border
                if col_idx != 5:
                    ws_status.cell(row=row_num, column=col_idx).font = font_normal

        # ----------------------------------------------------
        # SHEET 4: RIWAYAT PEMBAYARAN (KAS MASUK)
        # ----------------------------------------------------
        ws_pembayaran = wb.create_sheet(title="Riwayat Kas Masuk")
        ws_pembayaran.views.sheetView[0].showGridLines = True
        
        ws_pembayaran.append(["RIWAYAT PEMBAYARAN KAS MASUK"])
        ws_pembayaran["A1"].font = font_title
        ws_pembayaran.append([])
        
        headers_pay = ["ID", "Nama Siswa", "Kelas", "Bulan Iuran", "Nominal (Rp)", "Tanggal Bayar"]
        ws_pembayaran.append(headers_pay)
        ws_pembayaran.row_dimensions[3].height = 24
        
        for col_idx, h in enumerate(headers_pay, 1):
            cell = ws_pembayaran.cell(row=3, column=col_idx)
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = align_center
            cell.border = thin_border
            
        for p in pembayaran_list:
            ws_pembayaran.append([p.id, p.nama_siswa, p.kelas_siswa, p.bulan, p.nominal, p.tanggal])
            row_num = ws_pembayaran.max_row
            ws_pembayaran.row_dimensions[row_num].height = 18
            
            ws_pembayaran.cell(row=row_num, column=1).alignment = align_center
            ws_pembayaran.cell(row=row_num, column=2).alignment = align_left
            ws_pembayaran.cell(row=row_num, column=3).alignment = align_center
            ws_pembayaran.cell(row=row_num, column=4).alignment = align_center
            
            nom_cell = ws_pembayaran.cell(row=row_num, column=5)
            nom_cell.alignment = align_right
            nom_cell.number_format = 'Rp #,##0'
            
            ws_pembayaran.cell(row=row_num, column=6).alignment = align_center
            
            for col_idx in range(1, 7):
                ws_pembayaran.cell(row=row_num, column=col_idx).font = font_normal
                ws_pembayaran.cell(row=row_num, column=col_idx).border = thin_border

        # ----------------------------------------------------
        # SHEET 5: KAS KELUAR (PENGELUARAN)
        # ----------------------------------------------------
        ws_pengeluaran = wb.create_sheet(title="Kas Keluar")
        ws_pengeluaran.views.sheetView[0].showGridLines = True
        
        ws_pengeluaran.append(["RIWAYAT PENGELUARAN KAS KELAS"])
        ws_pengeluaran["A1"].font = font_title
        ws_pengeluaran.append([])
        
        headers_exp = ["ID", "Keterangan Pengeluaran", "Nominal (Rp)", "Tanggal Pengeluaran"]
        ws_pengeluaran.append(headers_exp)
        ws_pengeluaran.row_dimensions[3].height = 24
        
        for col_idx, h in enumerate(headers_exp, 1):
            cell = ws_pengeluaran.cell(row=3, column=col_idx)
            cell.font = font_header
            cell.fill = fill_secondary
            cell.alignment = align_center
            cell.border = thin_border
            
        for exp in pengeluaran_list:
            ws_pengeluaran.append([exp.id, exp.keterangan, exp.nominal, exp.tanggal])
            row_num = ws_pengeluaran.max_row
            ws_pengeluaran.row_dimensions[row_num].height = 18
            
            ws_pengeluaran.cell(row=row_num, column=1).alignment = align_center
            ws_pengeluaran.cell(row=row_num, column=2).alignment = align_left
            
            nom_cell = ws_pengeluaran.cell(row=row_num, column=3)
            nom_cell.alignment = align_right
            nom_cell.number_format = 'Rp #,##0'
            
            ws_pengeluaran.cell(row=row_num, column=4).alignment = align_center
            
            for col_idx in range(1, 5):
                ws_pengeluaran.cell(row=row_num, column=col_idx).font = font_normal
                ws_pengeluaran.cell(row=row_num, column=col_idx).border = thin_border

        # Auto-fit Column Widths across all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.row in [1, 2]: # Skip main titles
                        continue
                    val = str(cell.value or '')
                    # Format Rp cell string length approximation
                    if cell.number_format and 'Rp' in cell.number_format:
                        val = f"Rp {cell.value:,.0f}".replace(",", ".") if isinstance(cell.value, (int, float)) else str(cell.value)
                    if len(val) > max_len:
                        max_len = len(val)
                sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
        # Specific custom column width for sheet summary
        ws_summary.column_dimensions['B'].width = 28
        ws_summary.column_dimensions['C'].width = 24
        
        # Save Book
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Laporan_KasKu_{timestamp}.xlsx"
        filepath = os.path.join(settings.EXCEL_EXPORT_DIR, filename)
        wb.save(filepath)
        
        return filepath
